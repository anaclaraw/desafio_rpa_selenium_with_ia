import google.generativeai as genai
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import Workbook, load_workbook

def configure_webdriver(chromedriver_path):
    """Configura e inicializa o WebDriver do Selenium."""
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--headless")
    service = Service(chromedriver_path)
    return webdriver.Chrome(service=service, options=chrome_options)

def get_trending_videos(driver, category_url):
    """
    Acessa uma URL do YouTube e retorna os vídeos em alta na categoria.
    """
    try:
        driver.get(category_url)
        time.sleep(5)
        videos = driver.find_elements(By.ID, "meta")[:3]
        
        trending_videos = []
        for video in videos:
            title = video.find_element(By.ID, "video-title").get_attribute("title")
            link = video.find_element(By.ID, "video-title").get_attribute("href")
            canal = video.find_element(By.CSS_SELECTOR, "a.yt-simple-endpoint.style-scope.yt-formatted-string").text
            detalhes = video.find_elements(By.CSS_SELECTOR, "span.inline-metadata-item.style-scope.ytd-video-meta-block")
            visualizacoes = detalhes[0].text
            data_video = detalhes[1].text
            
            if title and link:                         
                trending_videos.append((title, link,visualizacoes,canal,data_video))
        
        
        return trending_videos
    except Exception as e:
        print(f"Erro ao buscar vídeos: {e}")
        return []

def save_to_excel(filepath, data):
    """
    Salva os dados fornecidos em um arquivo Excel.
    """
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Vídeos"
        sheet.append(["Categoria", "Título", "Visualizações","Canal","Data de Postagem","Link"])
        for category, videos in data.items():
            for title, link,visualizacoes,canal,data_video in videos:
                sheet.append([category, title, visualizacoes, canal, data_video,link])
        workbook.save(filepath)
        print("Arquivo Excel salvo com sucesso!")
    except Exception as e:
        print(f"Erro ao salvar no Excel: {e}")

def generate_ai_analysis(prompt, api_key, model_name="gemini-1.0-pro"):
    """
    Gera uma análise de IA com base no prompt e no modelo configurado.
    """
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(model_name=model_name)
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        print(f"Erro ao gerar análise com IA: {e}")
        return None

def main():
    """
    Função principal para executar o fluxo completo do programa.
    """
    # Configurações iniciais
    CHROMEDRIVER_PATH = "C:/Users/carol/AppData/Local/Microsoft/WindowsApps/chromedriver.exe"
    OUTPUT_FILE = "C:/Users/carol/Downloads/meu_arquivo.xlsx"
    GOOGLE_API_KEY = "#" 
    
    categories = {
        "Música": "https://www.youtube.com/feed/trending",
        "Jogos": "https://www.youtube.com/feed/trending?bp=4gIcGhpnYW1pbmdfY29ycHVzX21vc3RfcG9wdWxhcg%3D%3D",
        "Filmes": "https://www.youtube.com/feed/trending?bp=4gIKGgh0cmFpbGVycw%3D%3D"
    }

    # Inicializar o WebDriver
    driver = configure_webdriver(CHROMEDRIVER_PATH)
    
    try:
        # Coletar vídeos em alta por categoria
        trending_data = {}
        for category, url in categories.items():
            print(f"Coletando vídeos para a categoria: {category}")
            videos = get_trending_videos(driver, url)
            trending_data[category] = videos
        
        # Salvar os dados em um arquivo Excel
        save_to_excel(OUTPUT_FILE, trending_data)
        
        # Gerar análise com IA
        print("Gerando análise com IA...")
        workbook = load_workbook(OUTPUT_FILE)
        sheet = workbook.active
        video_titles = [row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[1]]
        prompt = f"Analise os seguintes títulos de vídeos em alta no YouTube: {video_titles}. Traga tendências e informações relevantes."
        analysis = generate_ai_analysis(prompt, GOOGLE_API_KEY)
        if analysis:
            print("Análise da IA:")
            print(analysis)
    
    except Exception as e:
        print(f"Erro na execução principal: {e}")
    
    finally:
        # Finalizar o WebDriver corretamente
        driver.quit()
        print("Execução finalizada.")

if __name__ == "__main__":
    main()
